#!/usr/bin/env python3
"""
逐个收集 Trae 任务的 trace 并回填到 trae_results_summary.xlsx

流程:
  1. 读取 Excel 的 task_name 列
  2. 过滤出 trace 为空（或 "null"）的行
  3. 逐个用 Trae 打开 tasks/autoX 文件夹
  4. 激活窗口后，两次快速尝试匹配 finish_all.png（阈值递减）
  5. 在其周围区域内找 finish.png
  6. 双击 finish.png，从剪贴板读取 trace
  7. 写回 Excel trace 列并保存

用法:
    python collect_trace.py
    python collect_trace.py --confidence 0.8
"""

import argparse
import ctypes
import io
import os
import subprocess
import sys
import time
from config_loader import get_path
from ctypes import wintypes

try:
    import openpyxl
    import pyautogui
    import pyperclip
except ImportError as e:
    print(f"[ERROR] 缺少依赖: {e}")
    print("请运行: pip install openpyxl pyautogui pyperclip opencv-python Pillow")
    sys.exit(1)

# Windows 终端默认 GBK，强制 stdout UTF-8
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 禁用 PyAutoGUI failsafe，防止鼠标移到左上角崩溃
pyautogui.FAILSAFE = False

# ===================== 配置 =====================
TRA_EXE = str(get_path("paths.trae_executable"))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "trae_results_summary.xlsx")
TASKS_DIR = os.path.join(BASE_DIR, "tasks")
# 使用相对路径避免中文绝对路径导致 OpenCV 编码问题（和 run.js / trae_helper.py 保持一致）
FINISH_ALL_IMAGE = "images/finish_all.png"
FINISH_IMAGE = "images/finish.png"
# ================================================

# ---------- Windows API ----------
user32 = ctypes.windll.user32

EnumWindows = user32.EnumWindows
EnumWindowsProc = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
GetWindowText = user32.GetWindowTextW
GetWindowTextLength = user32.GetWindowTextLengthW
IsWindowVisible = user32.IsWindowVisible
SetForegroundWindow = user32.SetForegroundWindow
ShowWindow = user32.ShowWindow
BringWindowToTop = user32.BringWindowToTop
SetWindowPos = user32.SetWindowPos

HWND_TOPMOST = -1
HWND_NOTOPMOST = -2
SWP_NOMOVE = 0x0002
SWP_NOSIZE = 0x0001
SWP_SHOWWINDOW = 0x0040
SW_MAXIMIZE = 3
WM_CLOSE = 0x0010


def find_window_by_title(keyword):
    """查找标题包含指定关键字的可见窗口，返回第一个匹配的 hwnd"""
    results = []

    def callback(hwnd, extra):
        if IsWindowVisible(hwnd):
            length = GetWindowTextLength(hwnd)
            if length > 0:
                buf = ctypes.create_unicode_buffer(length + 1)
                GetWindowText(hwnd, buf, length + 1)
                if keyword.lower() in buf.value.lower():
                    results.append(hwnd)
        return True

    EnumWindows(EnumWindowsProc(callback), 0)
    return results[0] if results else None


def activate_window(hwnd):
    """激活并最大化窗口（精简等待）"""
    SetWindowPos(hwnd, HWND_TOPMOST, 0, 0, 0, 0,
                 SWP_NOMOVE | SWP_NOSIZE | SWP_SHOWWINDOW)
    ShowWindow(hwnd, SW_MAXIMIZE)
    BringWindowToTop(hwnd)
    SetForegroundWindow(hwnd)
    time.sleep(1.0)
    SetWindowPos(hwnd, HWND_NOTOPMOST, 0, 0, 0, 0,
                 SWP_NOMOVE | SWP_NOSIZE | SWP_SHOWWINDOW)


def close_window(hwnd):
    """向窗口发送关闭消息"""
    user32.PostMessageW(hwnd, WM_CLOSE, 0, 0)


def open_trae_folder(folder_path):
    """用 Trae 打开指定文件夹，非阻塞"""
    print(f"  [Open] Trae -> {folder_path}")
    subprocess.Popen([TRA_EXE, "-n", folder_path],
                     stdout=subprocess.DEVNULL,
                     stderr=subprocess.DEVNULL)


def wait_for_finish_all(base_confidence):
    """
    两次快速尝试匹配 finish_all.png，阈值递减。
    第一次:  confidence=base_confidence      等 5s 后匹配
    第二次:  confidence=base_confidence-0.2   等 5s 后匹配
    两次总耗时约 10s，加上前后缓冲约 15s。
    成功返回 loc，失败返回 None，不空等。
    """
    confidences = [base_confidence, max(0.5, base_confidence - 0.2)]

    for attempt, conf in enumerate(confidences, 1):
        # 等待界面稳定（第一次额外等 2s 让 Trae 渲染）
        time.sleep(2 if attempt == 1 else 3)
        print(f"  [Try] finish_all.png 第 {attempt} 次尝试 (confidence={conf:.2f})")
        try:
            loc = pyautogui.locateOnScreen(FINISH_ALL_IMAGE, confidence=conf)
            if loc:
                print(f"  [Found] finish_all.png @ left={loc.left}, top={loc.top}, "
                      f"size=({loc.width}x{loc.height})")
                return loc
        except Exception:
            pass
        print(f"  [Miss] finish_all.png 第 {attempt} 次未匹配")

    return None


def find_finish_in_region(all_loc, base_confidence):
    """在 finish_all.png 周围扩大区域内搜索 finish.png，同样两次尝试"""
    if not all_loc:
        return None

    region = (
        max(0, all_loc.left - 50),
        max(0, all_loc.top - 30),
        all_loc.width + 100,
        all_loc.height + 60
    )

    confidences = [base_confidence, max(0.5, base_confidence - 0.2)]
    for attempt, conf in enumerate(confidences, 1):
        try:
            loc = pyautogui.locateCenterOnScreen(FINISH_IMAGE, confidence=conf, region=region)
            if loc:
                print(f"  [Found] finish.png @ ({loc.x}, {loc.y})")
                return loc
        except Exception:
            pass
        print(f"  [Miss] finish.png 第 {attempt} 次未匹配 (confidence={conf:.2f})")

    return None


def click_finish_and_read_trace(loc, verify_timeout):
    """双击 finish.png，轮询剪贴板读取 trace"""
    pyperclip.copy("")
    time.sleep(0.3)

    pyautogui.moveTo(loc.x, loc.y, duration=0.2)
    time.sleep(0.2)
    pyautogui.doubleClick(loc.x, loc.y)
    print(f"  [Click] 双击 finish.png ({loc.x}, {loc.y})")
    time.sleep(2)

    for v in range(1, verify_timeout + 1):
        time.sleep(1)
        content = pyperclip.paste()
        content_len = len(content) if content else 0
        if content and content_len > 10:
            return content
    return None


def get_trace_single(confidence, verify_timeout):
    """单窗口获取 trace，总耗时控制在约 15s"""
    # 1) 快速两次尝试匹配 finish_all.png
    all_loc = wait_for_finish_all(confidence)
    if not all_loc:
        print("  [Skip] finish_all.png 两次尝试均未匹配")
        return None

    # 2) 区域内两次尝试匹配 finish.png
    finish_loc = find_finish_in_region(all_loc, confidence)
    if not finish_loc:
        print("  [Skip] finish.png 两次尝试均未匹配")
        return None

    # 3) 点击并读取剪贴板
    trace = click_finish_and_read_trace(finish_loc, verify_timeout)
    return trace


def read_pending_tasks(excel_path):
    """从 Excel 读取 task_name 列表，过滤 trace 为空的行"""
    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel 文件不存在: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"[INFO] 表头: {headers}")

    try:
        task_col = headers.index("task_name") + 1
    except ValueError:
        print("[ERROR] Excel 中缺少 'task_name' 列")
        sys.exit(1)

    try:
        trace_col = headers.index("trace") + 1
    except ValueError:
        print("[ERROR] Excel 中缺少 'trace' 列")
        sys.exit(1)

    pending = []
    for row in range(2, ws.max_row + 1):
        task_name = ws.cell(row=row, column=task_col).value
        trace = ws.cell(row=row, column=trace_col).value

        if not task_name:
            continue

        trace_str = str(trace).strip() if trace else ""
        if trace_str and trace_str.lower() != "null":
            print(f"  [Skip] {task_name}: trace 已存在 (长度={len(trace_str)})")
            continue

        pending.append((row, task_name))

    wb.close()
    return pending


def save_trace_to_excel(excel_path, row, trace):
    """将 trace 写回指定行并保存"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    trace_col = headers.index("trace") + 1
    ws.cell(row=row, column=trace_col).value = trace
    wb.save(excel_path)
    wb.close()


def main():
    parser = argparse.ArgumentParser(description="逐个获取 Trae trace 并回填 Excel")
    parser.add_argument("--confidence", type=float, default=0.8, help="图像识别基础置信度 (默认 0.8，第二次会降 0.2)")
    parser.add_argument("--verify-timeout", type=int, default=5, help="点击后验证剪贴板轮询次数 (默认 5)")
    parser.add_argument("--keep-open", action="store_true", help="处理完后不关闭 Trae 窗口")
    parser.add_argument("--open-delay", type=int, default=12, help="打开文件夹后等待窗口加载的秒数 (默认 12)")
    args = parser.parse_args()

    os.chdir(BASE_DIR)
    print(f"[INFO] 工作目录已切换到: {os.getcwd()}")

    for img in (FINISH_ALL_IMAGE, FINISH_IMAGE):
        abs_path = os.path.abspath(img)
        if not os.path.exists(img):
            print(f"[ERROR] 图片文件缺失: {img} (绝对路径: {abs_path})")
            sys.exit(1)
        size = os.path.getsize(img)
        print(f"[INFO] 图片 OK: {img} (大小: {size} bytes, 绝对路径: {abs_path})")

    pending = read_pending_tasks(EXCEL_FILE)
    total = len(pending)
    print(f"\n[INFO] 共 {total} 个任务需要补录 trace\n")

    if total == 0:
        print("[INFO] 所有任务 trace 已填充完毕！")
        return

    success_count = 0
    fail_count = 0

    for idx, (row, task_name) in enumerate(pending, 1):
        print(f"\n{'='*60}")
        print(f"[{idx}/{total}] 处理任务: {task_name}")
        print(f"{'='*60}")

        folder = os.path.join(TASKS_DIR, task_name)
        if not os.path.exists(folder):
            print(f"  [ERROR] 文件夹不存在: {folder}")
            fail_count += 1
            continue

        # 关闭该任务已存在的窗口
        existing = find_window_by_title(task_name)
        if existing:
            print(f"  [Close] 关闭已存在的 {task_name} 窗口")
            close_window(existing)
            time.sleep(2)

        # 用 Trae 打开文件夹
        open_trae_folder(folder)
        time.sleep(args.open_delay)

        # 快速查找窗口句柄（最多 5s）
        hwnd = None
        for _ in range(5):
            hwnd = find_window_by_title(task_name)
            if hwnd:
                break
            time.sleep(1)

        if not hwnd:
            print(f"  [ERROR] 找不到 {task_name} 的窗口句柄")
            fail_count += 1
            continue

        # 激活窗口
        print(f"  [Activate] 激活窗口...")
        activate_window(hwnd)

        # 获取 trace（两次快速尝试，总耗时约 15s）
        trace = get_trace_single(
            confidence=args.confidence,
            verify_timeout=args.verify_timeout
        )

        if trace:
            save_trace_to_excel(EXCEL_FILE, row, trace)
            print(f"  [OK] trace 已回填 (长度={len(trace)})")
            success_count += 1
        else:
            print(f"  [FAIL] 未能获取 trace")
            fail_count += 1

        # 关闭窗口
        if not args.keep_open:
            print(f"  [Close] 关闭 {task_name} 窗口")
            close_window(hwnd)
            time.sleep(2)

    print(f"\n{'='*60}")
    print(f"处理完成！成功: {success_count} / 失败: {fail_count} / 总计: {total}")
    print(f"Excel 文件: {EXCEL_FILE}")
    print(f"{'='*60}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[INFO] 用户中断")
    except Exception as e:
        print(f"[ERROR] 脚本异常: {e}")
        import traceback
        traceback.print_exc()
