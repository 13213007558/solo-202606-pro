import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

ROUND_ORDER = {'第一轮': 1, '第二轮': 2, '第三轮': 3, '第四轮': 4, '第五轮': 5}

df = pd.read_excel('multi_round_results.xlsx')
original_count = len(df)

# 添加排序键并按 task_name + 轮次排序
df['_sort_key'] = df['多轮对话'].map(ROUND_ORDER)
df = df.sort_values(['task_name', '_sort_key']).reset_index(drop=True)

drop_indices = set()


def is_empty(val):
    """检查值是否为空（NaN、空字符串、纯空白、'nan'、'None'等）"""
    if pd.isna(val):
        return True
    s = str(val).strip().lower()
    return s == '' or s == 'nan' or s == 'none'


for task_name, group in df.groupby('task_name'):
    indices = group.index.tolist()
    sorted_rows = group.sort_values('_sort_key')

    first_round = sorted_rows[sorted_rows['多轮对话'] == '第一轮']

    # ========== 规则1：commit_id 相关检查 ==========

    # 第一轮 commit_id 为空 → 整组删除
    if first_round.empty or is_empty(first_round.iloc[0]['commit_id']):
        drop_indices.update(indices)
        continue

    # 组内 commit_id 有重复（忽略空值） → 整组删除
    commit_ids = group['commit_id'].dropna()
    commit_ids = commit_ids[commit_ids.astype(str).str.strip() != '']
    if commit_ids.duplicated().any():
        drop_indices.update(indices)
        continue

    # ========== 规则2：prompt 相关检查 ==========

    # 第一轮 prompt 为空 → 整组删除
    if is_empty(first_round.iloc[0]['prompt']):
        drop_indices.update(indices)
        continue

    # ========== 规则3：status 为"人工获取"检查 ==========

    # 第一轮 status 为"人工获取" → 整组删除
    if str(first_round.iloc[0].get('status', '')).strip() == '人工获取':
        drop_indices.update(indices)
        continue

    # ========== 逐轮检查（第2-5轮） ==========

    # Pass 1：检查第n轮 commit_id 或 prompt 为空 → 该轮及之后删除
    cutoff_round = None
    for _, row in sorted_rows.iterrows():
        round_num = ROUND_ORDER.get(row['多轮对话'], 99)
        if round_num < 2:
            continue
        if is_empty(row['commit_id']) or is_empty(row['prompt']):
            cutoff_round = round_num
            break

    if cutoff_round is not None:
        to_drop = [i for i in indices if df.loc[i, '_sort_key'] >= cutoff_round]
        drop_indices.update(to_drop)

    # Pass 2：检查 status 为"人工获取"且非第一轮 → 该轮及之后删除
    # （仅检查未被 Pass 1 删除的轮次）
    for _, row in sorted_rows.iterrows():
        round_num = ROUND_ORDER.get(row['多轮对话'], 99)
        if round_num < 2:
            continue
        if cutoff_round is not None and round_num >= cutoff_round:
            continue  # 已被 Pass 1 删除，跳过
        if str(row.get('status', '')).strip() == '人工获取':
            to_drop = [i for i in indices if df.loc[i, '_sort_key'] >= round_num]
            drop_indices.update(to_drop)
            break

# ========== 规则3：最终检查，没有"第一轮"的组整组删除 ==========
remaining_mask = ~df.index.isin(drop_indices)
remaining_df = df[remaining_mask]
for task_name, group in remaining_df.groupby('task_name'):
    if '第一轮' not in group['多轮对话'].values:
        drop_indices.update(group.index.tolist())

df = df.drop(index=drop_indices).drop(columns=['_sort_key']).reset_index(drop=True)

# 保存结果
output_path = 'multi_round_results_processed.xlsx'
df.to_excel(output_path, index=False)

# 标红 status 等于"人工获取"的单元格
wb = load_workbook(output_path)
ws = wb.active
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

status_col = None
for col_idx, cell in enumerate(ws[1], 1):
    if cell.value == 'status':
        status_col = col_idx
        break

if status_col:
    red_count = 0
    for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
        cell = row[0]
        if cell.value and str(cell.value).strip() == '人工获取':
            cell.fill = red_fill
            red_count += 1
    print(f"标红了 {red_count} 个 status='人工获取' 的单元格")

wb.save(output_path)

print(f"处理完成：原始 {original_count} 行，删除了 {len(drop_indices)} 行，剩余 {len(df)} 行")
print(f"结果已保存至 {output_path}")
