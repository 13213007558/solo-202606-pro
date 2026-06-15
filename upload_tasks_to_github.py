#!/usr/bin/env python3
"""
Upload task folders listed in trae_results_summary.xlsx to GitHub and write
the resulting repository location and commit ID back to the workbook.

Examples:
    python upload_tasks_to_github.py
    python upload_tasks_to_github.py --date 6.8 --branch main
    python upload_tasks_to_github.py --dry-run

Git authentication is handled by the local Git installation (normally Git
Credential Manager on Windows).
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import subprocess
import sys
import time
import winreg
from datetime import datetime
from pathlib import Path
from config_loader import get_config, get_path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: openpyxl. Install it with: pip install openpyxl")
    sys.exit(1)


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL = get_path("paths.results_excel", "trae_results_summary.xlsx")
DEFAULT_TASKS_DIR = get_path("paths.tasks_directory", "tasks")
DEFAULT_REPOSITORY = str(get_config("github.repository", ""))
DEFAULT_BRANCH = str(get_config("github.branch", "main"))
DEFAULT_REPOSITORY_DIR = get_path("github.repository_cache", ".github_upload/Solo")

TASK_NAME_HEADER = "task_name"
GITHUB_URL_HEADER = "github地址"
LOCATION_HEADER = "分支/文件夹"
COMMIT_ID_HEADER = "commit_id"

IGNORED_NAMES = {
    ".git",
    ".cache",
    ".mypy_cache",
    ".pytest_cache",
    ".vite",
    "__pycache__",
    "build",
    "coverage",
    "dist",
    "node_modules",
}


class UploadError(RuntimeError):
    pass


def normalize_proxy_url(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    if "://" not in value:
        value = f"http://{value}"
    return value


def detect_windows_proxy() -> str:
    try:
        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Internet Settings",
        ) as key:
            enabled = int(winreg.QueryValueEx(key, "ProxyEnable")[0])
            if not enabled:
                return ""
            proxy_server = str(winreg.QueryValueEx(key, "ProxyServer")[0]).strip()
    except (FileNotFoundError, OSError, ValueError):
        return ""

    if not proxy_server:
        return ""

    # Windows may store either "host:port" or
    # "http=host:port;https=host:port;socks=host:port".
    entries: dict[str, str] = {}
    if "=" in proxy_server:
        for item in proxy_server.split(";"):
            protocol, separator, address = item.partition("=")
            if separator and address.strip():
                entries[protocol.strip().lower()] = address.strip()
        proxy_server = (
            entries.get("https")
            or entries.get("http")
            or entries.get("socks")
            or ""
        )

    return normalize_proxy_url(proxy_server)


def build_git_environment(proxy: str) -> dict[str, str]:
    environment = os.environ.copy()
    if proxy:
        environment["HTTP_PROXY"] = proxy
        environment["HTTPS_PROXY"] = proxy
        environment["http_proxy"] = proxy
        environment["https_proxy"] = proxy
    return environment


def run_git(
    args: list[str],
    cwd: Path | None = None,
    check: bool = True,
    git_env: dict[str, str] | None = None,
) -> str:
    command = ["git", *args]
    result = subprocess.run(
        command,
        cwd=str(cwd) if cwd else None,
        env=git_env,
        text=True,
        encoding="utf-8",
        errors="replace",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    if check and result.returncode != 0:
        detail = result.stderr.strip() or result.stdout.strip()
        raise UploadError(f"Git command failed: {' '.join(command)}\n{detail}")
    return result.stdout.strip()


def run_git_with_retry(
    args: list[str],
    cwd: Path | None,
    attempts: int,
    retry_delay: int,
    operation: str,
    git_env: dict[str, str],
) -> str:
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return run_git(args, cwd=cwd, git_env=git_env)
        except UploadError as exc:
            last_error = exc
            if attempt >= attempts:
                break
            print(
                f"{operation} failed ({attempt}/{attempts}); "
                f"retrying in {retry_delay} seconds..."
            )
            time.sleep(retry_delay)
    raise UploadError(f"{operation} failed after {attempts} attempts.\n{last_error}")


def _is_non_fast_forward(error: UploadError) -> bool:
    msg = str(error).lower()
    return "non-fast-forward" in msg or "fetch first" in msg


def push_with_auto_sync(
    branch: str,
    cwd: Path,
    attempts: int,
    retry_delay: int,
    git_env: dict[str, str],
) -> str:
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return run_git(
                ["push", "-u", "origin", branch], cwd=cwd, git_env=git_env
            )
        except UploadError as exc:
            last_error = exc
            if _is_non_fast_forward(exc):
                print(
                    f"Push rejected: local branch is behind remote. "
                    f"Pulling with rebase (attempt {attempt}/{attempts})..."
                )
                try:
                    run_git(
                        ["pull", "--rebase", "origin", branch],
                        cwd=cwd,
                        git_env=git_env,
                    )
                    print("Rebase succeeded. Retrying push...")
                    continue
                except UploadError as pull_exc:
                    raise UploadError(
                        f"Automatic rebase failed. Manual resolution may be needed.\n"
                        f"{pull_exc}"
                    ) from pull_exc
            if attempt >= attempts:
                break
            print(
                f"Git push failed ({attempt}/{attempts}); "
                f"retrying in {retry_delay} seconds..."
            )
            time.sleep(retry_delay)
    raise UploadError(f"Git push failed after {attempts} attempts.\n{last_error}")


def github_web_url(repository: str) -> str:
    repository = repository.strip().rstrip("/")
    ssh_match = re.fullmatch(r"git@github\.com:(.+?)(?:\.git)?", repository)
    if ssh_match:
        return f"https://github.com/{ssh_match.group(1)}"

    https_match = re.fullmatch(
        r"https?://github\.com/(.+?)(?:\.git)?", repository, re.IGNORECASE
    )
    if https_match:
        return f"https://github.com/{https_match.group(1)}"

    raise UploadError(
        "Only GitHub HTTPS or SSH repository addresses are supported: "
        "https://github.com/owner/repo.git or git@github.com:owner/repo.git"
    )


def validate_date_label(value: str) -> str:
    value = value.strip()
    if not re.fullmatch(r"\d{1,2}\.\d{1,2}", value):
        raise UploadError("Date must use M.D format, for example 6.8")
    return value


def validate_folder_prefix(value: str) -> str:
    value = value.strip()
    if not re.fullmatch(r"[A-Za-z0-9_-]+", value):
        raise UploadError(
            "Folder prefix may contain only letters, numbers, underscores, and hyphens"
        )
    return value


def validate_task_name(value: object) -> str:
    task_name = str(value or "").strip()
    if not re.fullmatch(r"auto\d+", task_name, re.IGNORECASE):
        raise UploadError(f"Invalid task_name: {task_name!r}; expected auto followed by digits")
    return task_name


def workbook_lock_path(excel_path: Path) -> Path:
    return excel_path.with_name(f"~${excel_path.name}")


def ensure_workbook_available(excel_path: Path) -> None:
    if not excel_path.is_file():
        raise UploadError(f"Excel file not found: {excel_path}")
    lock_path = workbook_lock_path(excel_path)
    if lock_path.exists():
        raise UploadError(
            f"Excel appears to be open: {excel_path.name}. Close it before uploading."
        )


def ensure_headers(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column).value
        if value is not None:
            headers[str(value).strip()] = column

    if TASK_NAME_HEADER not in headers:
        raise UploadError(f"Excel is missing required column: {TASK_NAME_HEADER}")

    for header in (GITHUB_URL_HEADER, LOCATION_HEADER, COMMIT_ID_HEADER):
        if header not in headers:
            column = worksheet.max_column + 1
            worksheet.cell(row=1, column=column).value = header
            headers[header] = column

    return headers


def collect_tasks(worksheet, headers: dict[str, int], tasks_dir: Path):
    tasks = []
    seen_names: set[str] = set()

    for row in range(2, worksheet.max_row + 1):
        raw_name = worksheet.cell(row=row, column=headers[TASK_NAME_HEADER]).value
        if raw_name is None or not str(raw_name).strip():
            continue

        task_name = validate_task_name(raw_name)
        if task_name.lower() in seen_names:
            raise UploadError(f"Duplicate task_name in Excel: {task_name}")
        seen_names.add(task_name.lower())

        source = (tasks_dir / task_name).resolve()
        if source.parent != tasks_dir.resolve() or not source.is_dir():
            raise UploadError(f"Task folder not found: {source}")

        tasks.append({"row": row, "name": task_name, "source": source})

    if not tasks:
        raise UploadError("No task_name values were found in the Excel file")
    return tasks


def ignore_cache_files(_directory: str, names: list[str]) -> set[str]:
    ignored = {name for name in names if name in IGNORED_NAMES}
    ignored.update(
        name
        for name in names
        if name.endswith((".log", ".pyc", ".pyo")) or name.startswith("~$")
    )
    return ignored


def copy_task(source: Path, repository_dir: Path, folder_name: str) -> Path:
    destination = (repository_dir / folder_name).resolve()
    repository_root = repository_dir.resolve()
    if destination.parent != repository_root:
        raise UploadError(f"Unsafe destination path: {destination}")

    if destination.exists():
        shutil.rmtree(destination)
    shutil.copytree(source, destination, ignore=ignore_cache_files)
    return destination


def has_local_branch(repository_dir: Path, branch: str) -> bool:
    result = subprocess.run(
        ["git", "show-ref", "--verify", "--quiet", f"refs/heads/{branch}"],
        cwd=str(repository_dir),
    )
    return result.returncode == 0


def has_remote_branch(repository_dir: Path, branch: str) -> bool:
    result = subprocess.run(
        ["git", "show-ref", "--verify", "--quiet", f"refs/remotes/origin/{branch}"],
        cwd=str(repository_dir),
    )
    return result.returncode == 0


def checkout_upload_branch(repository_dir: Path, branch: str) -> None:
    if has_local_branch(repository_dir, branch):
        run_git(["checkout", branch], cwd=repository_dir)
    elif has_remote_branch(repository_dir, branch):
        run_git(["checkout", "-b", branch, f"origin/{branch}"], cwd=repository_dir)
    else:
        run_git(["checkout", "-B", branch], cwd=repository_dir)


def prepare_repository(
    repository: str,
    repository_dir: Path,
    branch: str,
    attempts: int,
    retry_delay: int,
    git_env: dict[str, str],
) -> None:
    git_dir = repository_dir / ".git"
    if git_dir.is_dir():
        print(f"Using persistent local repository: {repository_dir}")
        current_remote = run_git(
            ["remote", "get-url", "origin"],
            cwd=repository_dir,
            check=False,
            git_env=git_env,
        )
        if current_remote != repository:
            run_git(
                ["remote", "set-url", "origin", repository],
                cwd=repository_dir,
                git_env=git_env,
            )

        try:
            run_git_with_retry(
                ["fetch", "origin"],
                cwd=repository_dir,
                attempts=attempts,
                retry_delay=retry_delay,
                operation="Git fetch",
                git_env=git_env,
            )
        except UploadError as exc:
            print(f"WARNING: Could not fetch remote updates. Continuing locally.\n{exc}")
    else:
        if repository_dir.exists():
            if any(repository_dir.iterdir()):
                raise UploadError(
                    f"Repository cache exists but is not a Git repository: {repository_dir}"
                )
            repository_dir.rmdir()
        repository_dir.parent.mkdir(parents=True, exist_ok=True)
        print(f"Cloning repository into persistent cache: {repository_dir}")
        run_git_with_retry(
            ["clone", repository, str(repository_dir)],
            cwd=None,
            attempts=attempts,
            retry_delay=retry_delay,
            operation="Git clone",
            git_env=git_env,
        )

    checkout_upload_branch(repository_dir, branch)
    configure_commit_identity(repository_dir)


def configure_commit_identity(repository_dir: Path) -> None:
    name = run_git(["config", "user.name"], cwd=repository_dir, check=False)
    email = run_git(["config", "user.email"], cwd=repository_dir, check=False)
    if not name or not email:
        raise UploadError(
            "Git user identity is not configured. Run git config --global user.name "
            "and git config --global user.email first."
        )


def commit_task(repository_dir: Path, folder_name: str, task_name: str) -> str:
    run_git(["add", "--", folder_name], cwd=repository_dir)
    staged_changes = subprocess.run(
        ["git", "diff", "--cached", "--quiet", "--", folder_name],
        cwd=str(repository_dir),
    ).returncode

    if staged_changes == 0:
        existing_commit = run_git(
            ["log", "-1", "--format=%H", "--", folder_name],
            cwd=repository_dir,
            check=False,
        )
        if not existing_commit:
            raise UploadError(f"No changes or existing commit found for {folder_name}")
        print(f"  {task_name}: unchanged, using commit {existing_commit[:12]}")
        return existing_commit

    run_git(
        ["commit", "-m", f"Upload {task_name} as {folder_name}", "--", folder_name],
        cwd=repository_dir,
    )
    commit_id = run_git(["rev-parse", "HEAD"], cwd=repository_dir)
    print(f"  {task_name}: committed as {folder_name} ({commit_id[:12]})")
    return commit_id


def save_workbook_atomically(workbook, excel_path: Path) -> None:
    temporary_path = excel_path.with_name(f".{excel_path.stem}.uploading.xlsx")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, excel_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def upload(args) -> None:
    if args.retry_attempts < 1:
        raise UploadError("--retry-attempts must be at least 1")
    if args.retry_delay < 0:
        raise UploadError("--retry-delay cannot be negative")

    excel_path = Path(args.excel).resolve()
    tasks_dir = Path(args.tasks_dir).resolve()
    date_label = validate_date_label(args.date)
    folder_prefix = validate_folder_prefix(args.folder_prefix)
    web_url = github_web_url(args.repository)
    if args.no_proxy:
        proxy = ""
        print("Proxy: disabled by --no-proxy")
    elif args.proxy:
        proxy = normalize_proxy_url(args.proxy)
        print(f"Proxy: {proxy} (command line)")
    else:
        proxy = detect_windows_proxy()
        if proxy:
            print(f"Proxy: {proxy} (Windows system proxy)")
        else:
            print("Proxy: not detected; Git will use its existing configuration")
    git_env = build_git_environment(proxy)

    ensure_workbook_available(excel_path)
    if not tasks_dir.is_dir():
        raise UploadError(f"Tasks directory not found: {tasks_dir}")

    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active
    headers = ensure_headers(worksheet)
    tasks = collect_tasks(worksheet, headers, tasks_dir)

    print(f"Excel: {excel_path}")
    print(f"Repository: {args.repository}")
    print(f"Branch: {args.branch}")
    print(f"Tasks: {len(tasks)}")

    if args.dry_run:
        for task in tasks:
            folder_name = f"{folder_prefix}_{date_label}_{task['name']}"
            print(f"  {task['name']} -> {args.branch}/{folder_name}")
        print("Dry run complete; no files, Git repository, or Excel data were changed.")
        return

    repository_dir = Path(args.repository_dir).resolve()
    prepare_repository(
        repository=args.repository,
        repository_dir=repository_dir,
        branch=args.branch,
        attempts=args.retry_attempts,
        retry_delay=args.retry_delay,
        git_env=git_env,
    )

    results = []
    for task in tasks:
        folder_name = f"{folder_prefix}_{date_label}_{task['name']}"
        copy_task(task["source"], repository_dir, folder_name)
        commit_id = commit_task(
            repository_dir, folder_name=folder_name, task_name=task["name"]
        )
        results.append(
            {
                **task,
                "folder": folder_name,
                "commit_id": commit_id,
                "github_url": web_url,
            }
        )

    print("Pushing commits...")
    try:
        push_with_auto_sync(
            branch=args.branch,
            cwd=repository_dir,
            attempts=args.retry_attempts,
            retry_delay=args.retry_delay,
            git_env=git_env,
        )
    except UploadError as exc:
        raise UploadError(
            f"{exc}\nLocal commits were preserved in: {repository_dir}\n"
            "Run the same command again after the network connection is restored."
        ) from exc

    # Recheck immediately before replacing the workbook.
    ensure_workbook_available(excel_path)
    for result in results:
        row = result["row"]
        worksheet.cell(row=row, column=headers[GITHUB_URL_HEADER]).value = result[
            "github_url"
        ]
        worksheet.cell(row=row, column=headers[LOCATION_HEADER]).value = (
            f"{args.branch}/{result['folder']}"
        )
        worksheet.cell(row=row, column=headers[COMMIT_ID_HEADER]).value = result[
            "commit_id"
        ]

    save_workbook_atomically(workbook, excel_path)

    print(f"Upload complete. Excel updated: {excel_path}")


def build_parser() -> argparse.ArgumentParser:
    now = datetime.now()
    parser = argparse.ArgumentParser(
        description="Upload task folders to GitHub and update trae_results_summary.xlsx"
    )
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Path to the xlsx file")
    parser.add_argument(
        "--tasks-dir", default=str(DEFAULT_TASKS_DIR), help="Directory containing auto* folders"
    )
    parser.add_argument("--repository", default=DEFAULT_REPOSITORY, help="Git repository URL")
    parser.add_argument(
        "--repository-dir",
        default=str(DEFAULT_REPOSITORY_DIR),
        help="Persistent local Git repository used for retry and recovery",
    )
    parser.add_argument("--branch", default=DEFAULT_BRANCH, help="Target Git branch")
    parser.add_argument(
        "--date",
        default=f"{now.month}.{now.day}",
        help="Folder date prefix in M.D format; defaults to today",
    )
    parser.add_argument(
        "--folder-prefix",
        default=str(get_config("github.folder_prefix", "cp1")),
        help="Folder prefix placed before the date; defaults to cp1",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and display planned uploads without changing anything",
    )
    parser.add_argument(
        "--retry-attempts",
        type=int,
        default=int(get_config("github.retry_attempts", 3)),
        help="Number of attempts for clone, fetch, and push",
    )
    parser.add_argument(
        "--retry-delay",
        type=int,
        default=int(get_config("github.retry_delay_seconds", 10)),
        help="Seconds to wait between network retries",
    )
    proxy_group = parser.add_mutually_exclusive_group()
    proxy_group.add_argument(
        "--proxy",
        help="Proxy for this run, for example http://127.0.0.1:7890",
    )
    proxy_group.add_argument(
        "--no-proxy",
        action="store_true",
        help="Do not auto-detect or inject the Windows system proxy",
    )
    return parser


def main() -> int:
    try:
        upload(build_parser().parse_args())
        return 0
    except (UploadError, OSError, PermissionError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
