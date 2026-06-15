#!/usr/bin/env python3
"""
遍历 auto1 ~ autoN 目录，读取 result_autoN.json，
提取指定字段并导出到 Excel。
"""

import io
import json
import os
import re
import sys

if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TASKS_DIR = os.path.join(BASE_DIR, "tasks")
OUTPUT_FILE = os.path.join(BASE_DIR, "trae_results_summary.xlsx")


def clean_text(text):
    """去除换行符、非法控制字符和多余空格"""
    if text is None:
        return ""
    text = str(text)
    # 去除 Excel 不允许的控制字符（保留 \t \n \r）
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    text = text.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def main():
    # 收集所有 auto 目录
    auto_dirs = []
    search_dir = TASKS_DIR if os.path.isdir(TASKS_DIR) else BASE_DIR
    for entry in os.listdir(search_dir):
        full_path = os.path.join(search_dir, entry)
        if os.path.isdir(full_path) and re.match(r"^auto\d+$", entry):
            auto_dirs.append(entry)

    auto_dirs.sort(key=lambda x: int(re.search(r"\d+", x).group()))

    if not auto_dirs:
        print("未找到 auto 目录")
        sys.exit(0)

    print(f"找到 {len(auto_dirs)} 个目录: {', '.join(auto_dirs)}")

    # Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "task_name",
        "task_type",
        "business_domain",
        "modification_scope",
        "prompt",
        "session_id_full_id",
        "trace",
    ]
    ws.append(headers)

    total = 0
    for d in auto_dirs:
        result_file = os.path.join(search_dir, d, f"result_{d}.json")
        if not os.path.exists(result_file):
            print(f"  ⚠️ 跳过 {d}: {result_file} 不存在")
            continue

        with open(result_file, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError as e:
                print(f"  ❌ 跳过 {d}: JSON 解析失败 - {e}")
                continue

        task_name = clean_text(data.get("task_name", ""))
        task_type = clean_text(data.get("task_type", ""))
        business_domain = clean_text(data.get("business_domain", ""))
        modification_scope = clean_text(data.get("modification_scope", ""))
        prompt = clean_text(data.get("prompt", ""))

        session_id = data.get("session_id")
        if isinstance(session_id, dict):
            full_id = clean_text(session_id.get("full_id", ""))
        else:
            full_id = ""

        trace = clean_text(data.get("trace", ""))

        ws.append([task_name, task_type, business_domain, modification_scope, prompt, full_id, trace])
        total += 1
        print(f"  ✅ 已导入 {d}")

    # 设置列宽和自动换行
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 60

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT_FILE)
    print(f"\n📊 导出完成: {OUTPUT_FILE}")
    print(f"   共 {total} 条记录")


if __name__ == "__main__":
    main()
