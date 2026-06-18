#!/usr/bin/env python3
"""
多轮对话辅助脚本
提供：git_upload, clean_trace, fill_dissatisfaction, generate_next_prompt, export_xlsx
"""

import argparse
import io
import json
import os
import random
import re
import shutil
import subprocess
import sys
import time

from config_loader import BASE_DIR as CONFIG_BASE_DIR, get_config, get_deepseek_config, get_path

if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE_DIR = str(CONFIG_BASE_DIR)
TASKS_DIR = str(get_path('paths.tasks_directory', 'tasks'))

# DeepSeek API 配置
API_KEY, BASE_URL, MODEL = get_deepseek_config()

ROUND_NAMES = ["", "第一轮", "第二轮", "第三轮", "第四轮", "第五轮",
               "第六轮", "第七轮", "第八轮", "第九轮", "第十轮"]


def get_ai_client():
    api_key, base_url, _ = get_deepseek_config(required=True)
    try:
        from openai import OpenAI
        return OpenAI(api_key=api_key, base_url=base_url)
    except ImportError:
        print(json.dumps({"success": False, "error": "Missing openai package. Run: pip install openai"}))
        sys.exit(1)


def load_tasks(tasks_json_str):
    return json.loads(tasks_json_str)


def load_result(task_path, task_name):
    result_file = os.path.join(task_path, f'result_{task_name}.json')
    if os.path.exists(result_file):
        with open(result_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"task_name": task_name, "rounds": []}


def save_result(task_path, task_name, data):
    result_file = os.path.join(task_path, f'result_{task_name}.json')
    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_round_data(result, round_num):
    for r in result.get('rounds', []):
        if r.get('round') == round_num:
            return r
    return None


def set_round_data(result, round_num, data):
    if 'rounds' not in result:
        result['rounds'] = []
    for i, r in enumerate(result['rounds']):
        if r.get('round') == round_num:
            result['rounds'][i] = data
            return
    result['rounds'].append(data)


# ============================================================
# ACTION: git_upload
# ============================================================
def action_git_upload(args):
    tasks = load_tasks(args.tasks_json)
    repo = args.repo
    branch = args.branch
    prefix = args.prefix
    date_str = args.date or time.strftime('%m.%d').lstrip('0').replace('.0', '.')
    round_num = args.round

    clone_dir = str(get_path('github.repository_cache', '.github_upload/Solo'))

    def run_git(cmd, cwd=None, timeout=30):
        try:
            result = subprocess.run(
                cmd, cwd=cwd or clone_dir, capture_output=True,
                text=True, timeout=timeout, encoding='utf-8', errors='replace'
            )
            return result.returncode == 0, result.stdout, result.stderr
        except Exception as e:
            return False, '', str(e)

    # 准备仓库
    if os.path.exists(os.path.join(clone_dir, '.git')):
        ok, _, err = run_git(['git', 'remote', 'set-url', 'origin', repo])
        if not ok:
            return {"success": False, "error": f"Could not update repository remote: {err}"}
        run_git(['git', 'fetch', 'origin'], timeout=60)
    else:
        os.makedirs(os.path.dirname(clone_dir), exist_ok=True)
        ok, out, err = run_git(
            ['git', 'clone', repo, clone_dir], cwd=BASE_DIR, timeout=120
        )
        if not ok:
            return {"success": False, "error": f"Clone failed: {err}"}

    def has_head_commit():
        ok, _, _ = run_git(['git', 'rev-parse', '--verify', 'HEAD'])
        return ok

    def checkout_upload_branch():
        ok, _, err = run_git(['git', 'checkout', branch])
        if ok:
            return True, ''

        remote_ok, _, _ = run_git(
            ['git', 'rev-parse', '--verify', f'origin/{branch}']
        )
        if remote_ok:
            ok, _, err = run_git(
                ['git', 'checkout', '-B', branch, f'origin/{branch}']
            )
            return (True, '') if ok else (False, err)

        ok, _, err = run_git(['git', 'checkout', '-B', branch])
        if ok:
            return True, ''
        return False, err

    ok, err = checkout_upload_branch()
    if not ok:
        return {"success": False, "error": f"Checkout failed: {err}"}
    if has_head_commit():
        ok, _, err = run_git(['git', 'reset'])
        if not ok:
            return {"success": False, "error": f"Could not reset Git index: {err}"}

    ignore_patterns = {'.git', 'node_modules', '__pycache__', 'dist', 'build', '.cache'}
    ignore_extensions = {'.log', '.pyc', '.pyo'}
    ignore_prefixes = ('prompt_', 'result_auto')

    def ignore_func(directory, contents):
        ignored = set()
        for item in contents:
            if item in ignore_patterns:
                ignored.add(item)
            elif any(item.startswith(p) for p in ignore_prefixes):
                ignored.add(item)
            elif any(item.endswith(ext) for ext in ignore_extensions):
                ignored.add(item)
            elif item.startswith('~$'):
                ignored.add(item)
        return ignored

    results = []
    failures = []

    def add_failure(task_name, error):
        failures.append({"task_name": task_name, "error": error})

    for task in tasks:
        task_name = task['name']
        task_path = task['path']
        folder_name = f"{prefix}_{date_str}_{task_name}"
        dest_path = os.path.join(clone_dir, folder_name)

        if os.path.exists(dest_path):
            shutil.rmtree(dest_path)
        shutil.copytree(task_path, dest_path, ignore=ignore_func)

        ok, _, err = run_git(['git', 'add', '--', folder_name])
        if not ok:
            add_failure(task_name, f"git add failed: {err.strip()}")
            continue

        ok, _, _ = run_git(['git', 'diff', '--cached', '--quiet', '--', folder_name])
        if ok:
            head_ok, head_out, head_err = run_git(
                ['git', 'ls-tree', '-r', '--name-only', 'HEAD', '--', folder_name]
            )
            head_paths = [line.strip() for line in head_out.splitlines() if line.strip()]
            if not head_ok or not any(
                path == folder_name or path.startswith(folder_name + '/')
                for path in head_paths
            ):
                add_failure(
                    task_name,
                    (
                        f"No uploadable tracked files for {folder_name}; "
                        f"empty commit would not contain the folder. {head_err.strip()}"
                    ).strip()
                )
                continue

            msg = (
                f"Round {round_num}: Record unchanged {task_name} "
                f"as {folder_name}"
            )
            commit_ok, _, commit_err = run_git(
                ['git', 'commit', '--allow-empty', '-m', msg]
            )
            if not commit_ok:
                add_failure(
                    task_name,
                    f"Empty commit failed for unchanged task: {commit_err.strip()}"
                )
                continue
            ok2, out2, err2 = run_git(['git', 'rev-parse', 'HEAD'])
            commit_id = out2.strip() if ok2 else ''
            if not commit_id:
                add_failure(task_name, f"Could not read commit ID: {err2.strip()}")
                continue
        else:
            msg = f"Round {round_num}: Upload {task_name} as {folder_name}"
            before_ok, before_out, _ = run_git(['git', 'rev-parse', 'HEAD'])
            before_commit = before_out.strip() if before_ok else ''
            commit_ok, _, commit_err = run_git(
                ['git', 'commit', '-m', msg]
            )
            if not commit_ok:
                add_failure(task_name, f"Commit failed: {commit_err.strip()}")
                continue
            ok2, out2, err2 = run_git(['git', 'rev-parse', 'HEAD'])
            commit_id = out2.strip() if ok2 else ''
            if not commit_id or commit_id == before_commit:
                add_failure(
                    task_name,
                    f"Git did not create a new commit: {err2.strip()}"
                )
                continue

        verify_ok, verify_out, verify_err = run_git(
            ['git', 'ls-tree', '-r', '--name-only', commit_id, '--', folder_name]
        )
        tree_paths = [line.strip() for line in verify_out.splitlines() if line.strip()]
        if not verify_ok or not any(
            path == folder_name or path.startswith(folder_name + '/')
            for path in tree_paths
        ):
            add_failure(
                task_name,
                (
                    f"Commit verification failed; commit {commit_id} "
                    f"does not contain {folder_name}. {verify_err.strip()}"
                ).strip()
            )
            continue

        github_url = repo.replace('.git', '')
        results.append({
            "task_name": task_name,
            "github_url": github_url,
            "branch_folder": f"{branch}/{folder_name}",
            "commit_id": commit_id
        })

    commit_owners = {}
    for result in results:
        commit_id = result["commit_id"]
        previous_task = commit_owners.get(commit_id)
        if previous_task:
            add_failure(
                result['task_name'],
                (
                    f"Duplicate commit ID detected in the same batch: "
                    f"{previous_task} and {result['task_name']} both use {commit_id}"
                )
            )
        commit_owners[commit_id] = result["task_name"]

    if not results:
        return {
            "success": False,
            "error": "No tasks were uploaded",
            "data": {"results": results, "failures": failures}
        }

    # Push
    push_ok = False
    push_error = ""
    for attempt in range(int(get_config('github.retry_attempts', 3))):
        ok, out, err = run_git(['git', 'push', '-u', 'origin', branch], timeout=60)
        if ok:
            push_ok = True
            break
        push_error = err
        if 'non-fast-forward' in err or 'rejected' in err:
            run_git(['git', 'pull', '--rebase', 'origin', branch], timeout=60)
        else:
            time.sleep(3)
    if not push_ok:
        return {
            "success": False,
            "error": f"Git push failed: {push_error.strip()}",
            "data": {"results": results, "failures": failures}
        }

    # 保存到 result 文件
    for r in results:
        task_name = r['task_name']
        task_path = os.path.join(TASKS_DIR, task_name)
        result_data = load_result(task_path, task_name)
        rd = get_round_data(result_data, round_num)
        if rd:
            rd['github_url'] = r['github_url']
            rd['branch_folder'] = r['branch_folder']
            rd['commit_id'] = r['commit_id']
            set_round_data(result_data, round_num, rd)
            save_result(task_path, task_name, result_data)

    return {"success": True, "data": {"results": results, "failures": failures}}


# ============================================================
# ACTION: clean_trace
# ============================================================
def action_clean_trace(args):
    tasks = load_tasks(args.tasks_json)
    round_num = args.round
    return {"success": True, "data": {"cleaned": 0, "preserved": len(tasks), "round": round_num}}


# ============================================================
# ACTION: fill_dissatisfaction
# ============================================================
DISSATISFACTION_SYSTEM_PROMPT = """你是一位经验丰富的技术负责人，正在审核AI交付的工作产物。请根据用户的需求、AI的执行过程以及当前项目实际交付的代码文件，写出具体的不满意原因。

输出格式要求：
产物不满意：
1. [具体问题描述，包含技术细节]
2. ...

过程不满意：
1. [具体问题描述]
2. ...

要求：
- 必须结合下面提供的"当前项目代码"进行逐文件审查，指出代码中具体的错误、遗漏或与需求不符的实现
- 每条原因必须具体、可操作，包含技术细节（如文件名、函数名、变量名或代码逻辑问题）
- 不要写笼统的评价，必须指出具体的代码问题或流程缺陷
- 不要使用AI/模型/智能体等元语言
- 语气像经验丰富的技术负责人在审查下属工作
- 产物不满意至少3条，过程不满意至少2条"""


def action_fill_dissatisfaction(args):
    tasks = load_tasks(args.tasks_json)
    round_num = args.round
    client = get_ai_client()
    filled = 0

    for task in tasks:
        task_name = task['name']
        task_path = task['path']
        result_data = load_result(task_path, task_name)
        rd = get_round_data(result_data, round_num)
        if not rd:
            continue
        if rd.get('不满意原因'):
            continue

        prompt = rd.get('prompt', '')
        trace = rd.get('trace', '')
        if not prompt and not trace:
            continue

        code_preview = scan_code_files(task_path, max_files=8, max_chars=5000)

        user_msg = f"""【用户需求 Prompt】：
{prompt or '(无)'}

【AI执行过程 Trace】：
{trace or '(无)'}

【当前项目代码（部分）】：
{code_preview}

请根据以上内容，并结合"当前项目代码"逐文件审查，严格按照格式要求写出产物不满意原因和过程不满意原因。"""

        for attempt in range(3):
            try:
                response = client.chat.completions.create(
                    model=MODEL,
                    messages=[
                        {"role": "system", "content": DISSATISFACTION_SYSTEM_PROMPT},
                        {"role": "user", "content": user_msg}
                    ],
                    temperature=0.3,
                    max_tokens=4096
                )
                content = response.choices[0].message.content
                if "产物不满意" in content and "过程不满意" in content:
                    rd['不满意原因'] = content
                    set_round_data(result_data, round_num, rd)
                    save_result(task_path, task_name, result_data)
                    filled += 1
                    break
            except Exception as e:
                print(f"  API error for {task_name}: {e}", file=sys.stderr)
                time.sleep(2 * (attempt + 1))

        time.sleep(1)

    return {"success": True, "data": {"filled": filled}}


# ============================================================
# ACTION: generate_next_prompt
# ============================================================
NEXT_PROMPT_SYSTEM = """你是一个软件开发者，正在使用AI编程工具开发项目，现在需要对已有代码提出修改要求。

你需要根据以下信息生成一段自然的修改指令：
1. 当前代码存在的问题（产物问题和过程问题）
2. 代码执行日志
3. 项目当前的代码文件
4. 第一轮的原始需求（用来判断哪些功能是原需求要求的）

生成的提示词要求：
- 格式要像真实用户在对话中自然地提出修改要求
- 不要太AI化，语气自然直接
- 严禁出现"上一轮"、"本轮"、"下一轮"、"前一次"、"这一次"等指代轮次的词语，直接描述要修改什么
- 要具体指出需要修改什么，不要笼统
- 可以包含多个修改点
- 不要使用markdown格式标记（如**、##、「」等）
- 直接描述需求，不要写"请帮我"之类的开头
- 字数控制在200-500字

判断修改类型的规则（重要）：
- Bug修复：代码报错、功能缺失（原始需求要求但没实现的）、运行异常、页面崩溃、功能不正常等
- Feature迭代：在已有功能基础上优化、增强、调整样式、提升性能等
- 如果问题中涉及"缺少xxx页面"、"没有实现xxx功能"且这些是原始需求中要求的，这属于Bug修复而不是Feature迭代
- Feature迭代必须与原始项目需求强相关，不能凭空添加原始需求中完全没提到的新功能模块"""

TASK_TYPE_BUG_FIX = "Bug修复"
TASK_TYPE_FEATURE_ITERATION = "Feature迭代"


def get_task_type_probabilities():
    probabilities = get_config(
        "automation.subsequent_round_task_type_probabilities", {}
    )
    if not isinstance(probabilities, dict):
        raise ValueError(
            "automation.subsequent_round_task_type_probabilities must be an object"
        )

    bug_probability = float(probabilities.get("bug_fix", 0.6))
    feature_probability = float(probabilities.get("feature_iteration", 0.4))
    if not 0 <= bug_probability <= 1 or not 0 <= feature_probability <= 1:
        raise ValueError("Task type probabilities must be between 0 and 1")
    if abs((bug_probability + feature_probability) - 1.0) > 1e-9:
        raise ValueError(
            "bug_fix and feature_iteration probabilities must add up to 1"
        )
    return bug_probability, feature_probability


def choose_task_type():
    bug_probability, _ = get_task_type_probabilities()
    return (
        TASK_TYPE_BUG_FIX
        if random.random() < bug_probability
        else TASK_TYPE_FEATURE_ITERATION
    )


def build_task_type_instruction(task_type):
    if task_type == TASK_TYPE_BUG_FIX:
        return """本次任务类型已经确定为“Bug修复”。
只允许生成修复已有错误、异常、缺失功能或不符合原始需求行为的提示词。
不得新增原始需求之外的功能，不得把任务写成产品增强或功能扩展。
提示词必须以“修复以下问题：”开头。"""

    return """本次任务类型已经确定为“Feature迭代”。
只允许基于现有项目和原始需求生成优化、增强或合理扩展的提示词。
不得把已有报错、功能缺失、运行异常或不符合原始需求的问题包装成Feature。
提示词必须以“在现有功能基础上进行迭代：”开头。"""


def prompt_matches_task_type(task_type, prompt_content):
    prompt_content = prompt_content.strip()
    if task_type == TASK_TYPE_BUG_FIX:
        return prompt_content.startswith("修复以下问题：")
    if not prompt_content.startswith("在现有功能基础上进行迭代："):
        return False

    bug_semantic_markers = (
        "修复",
        "报错",
        "错误",
        "异常",
        "崩溃",
        "缺失",
        "未实现",
        "不生效",
        "失效",
        "不正常",
    )
    return not any(marker in prompt_content for marker in bug_semantic_markers)


def scan_code_files(task_path, max_files=5, max_chars=3000):
    code_extensions = {'.js', '.ts', '.tsx', '.jsx', '.py', '.html', '.css', '.vue', '.json'}
    ignore_dirs = {'node_modules', '.git', '__pycache__', 'dist', 'build', '.cache'}
    code_snippets = []
    total_chars = 0

    for root, dirs, files in os.walk(task_path):
        dirs[:] = [d for d in dirs if d not in ignore_dirs]
        for f in sorted(files):
            if len(code_snippets) >= max_files:
                break
            ext = os.path.splitext(f)[1].lower()
            if ext not in code_extensions:
                continue
            if f.startswith('prompt_') or f.startswith('result_'):
                continue
            filepath = os.path.join(root, f)
            relpath = os.path.relpath(filepath, task_path)
            try:
                with open(filepath, 'r', encoding='utf-8', errors='replace') as fh:
                    content = fh.read(max_chars - total_chars)
                code_snippets.append(f"--- {relpath} ---\n{content}")
                total_chars += len(content)
                if total_chars >= max_chars:
                    break
            except Exception:
                continue

    return '\n\n'.join(code_snippets) if code_snippets else '(无代码文件)'


def action_generate_next_prompt(args):
    tasks = load_tasks(args.tasks_json)
    round_num = args.round
    client = get_ai_client()
    generated = 0

    for task in tasks:
        task_name = task['name']
        task_path = task['path']
        result_data = load_result(task_path, task_name)

        prev_round = get_round_data(result_data, round_num - 1)
        if not prev_round:
            continue

        dissatisfaction = prev_round.get('不满意原因', '')
        trace = prev_round.get('trace', '')
        code_preview = scan_code_files(task_path)

        first_round = get_round_data(result_data, 1)
        business_domain = first_round.get('business_domain', '') if first_round else ''
        modification_scope = first_round.get('modification_scope', '') if first_round else ''
        original_prompt = first_round.get('prompt', '') if first_round else ''

        task_type = choose_task_type()
        task_type_instruction = build_task_type_instruction(task_type)
        if task_type == TASK_TYPE_BUG_FIX:
            task_context = f"""当前存在的问题：
{dissatisfaction or '(无)'}

执行日志（部分）：
{(trace or '(无)')[:2000]}"""
        else:
            task_context = """请从原始需求和当前代码中选择一个与项目强相关的优化或增强方向。
不要使用“不满意原因”中的缺陷作为Feature内容，也不要要求修复报错、缺失功能或异常行为。"""

        user_msg = f"""{task_type_instruction}

{task_context}

第一轮的原始需求：
{(original_prompt or '(无)')[:1000]}

当前项目代码（部分）：
{code_preview}

请严格按照已经确定的任务类型，写一段自然的修改指令。"""

        for attempt in range(3):
            try:
                response = client.chat.completions.create(
                    model=MODEL,
                    messages=[
                        {"role": "system", "content": NEXT_PROMPT_SYSTEM},
                        {"role": "user", "content": user_msg}
                    ],
                    temperature=0.5,
                    max_tokens=2048
                )
                prompt_content = response.choices[0].message.content.strip()
                if len(prompt_content) > 50 and prompt_matches_task_type(
                    task_type, prompt_content
                ):
                    prompt_data = {
                        "任务类型": task_type,
                        "业务领域": business_domain,
                        "修改范围": modification_scope,
                        "提示词内容": prompt_content
                    }

                    prompt_file = os.path.join(task_path, f'prompt_{round_num}.json')
                    with open(prompt_file, 'w', encoding='utf-8') as f:
                        json.dump(prompt_data, f, ensure_ascii=False, indent=2)

                    generated += 1
                    break
                print(
                    f"  Prompt type mismatch for {task_name}: expected {task_type}; retrying",
                    file=sys.stderr,
                )
            except Exception as e:
                print(f"  API error for {task_name}: {e}", file=sys.stderr)
                time.sleep(2 * (attempt + 1))

        time.sleep(1)

    return {"success": True, "data": {"generated": generated}}


# ============================================================
# ACTION: export_xlsx
# ============================================================
def clean_text(text):
    if not text:
        return ''
    text = str(text)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    text = re.sub(r'\s{2,}', ' ', text)
    return text.strip()


def action_export_xlsx(args):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
    except ImportError:
        return {"success": False, "error": "Missing openpyxl. Run: pip install openpyxl"}

    tasks = load_tasks(args.tasks_json)
    total_rounds = args.rounds

    headers = [
        'session_id_full_id', '多轮对话', 'prompt', 'task_type',
        'business_domain', 'modification_scope',
        '任务是否完成', '产物及过程是否满意', '不满意原因',
        'github地址', 'commit_id', '分支/文件夹', 'trace', 'status', 'task_name'
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    row = 2
    for task in tasks:
        task_name = task['name']
        task_path = task['path']
        result_data = load_result(task_path, task_name)

        for round_num in range(1, total_rounds + 1):
            rd = get_round_data(result_data, round_num)
            if not rd:
                rd = {}

            session_id = ''
            sid = rd.get('session_id')
            if isinstance(sid, dict):
                session_id = sid.get('full_id', '') or ''
            elif isinstance(sid, str):
                session_id = sid

            round_label = ROUND_NAMES[round_num] if round_num < len(ROUND_NAMES) else f"第{round_num}轮"

            status_parts = []
            if not session_id:
                status_parts.append('session_id')
            if not rd.get('trace') or rd.get('trace') == 'null':
                status_parts.append('trace')
            if not rd.get('github_url'):
                status_parts.append('git')
            status = '人工获取' if status_parts else 'ok'

            is_satisfied = rd.get('satisfied', False)
            task_complete = rd.get('任务是否完成', '完成了任务' if is_satisfied else '未完成任务')
            is_happy = rd.get('产物及过程是否满意', '满意' if is_satisfied else '不满意')

            values = [
                clean_text(session_id),
                round_label,
                clean_text(rd.get('prompt', '')),
                clean_text(rd.get('task_type', '')),
                clean_text(rd.get('business_domain', '')),
                clean_text(rd.get('modification_scope', '')),
                task_complete,
                is_happy,
                clean_text(rd.get('不满意原因', '')),
                clean_text(rd.get('github_url', '')),
                clean_text(rd.get('commit_id', '')),
                clean_text(rd.get('branch_folder', '')),
                clean_text(rd.get('trace', '')),
                status,
                task_name
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

    col_widths = [40, 10, 60, 15, 15, 18, 12, 15, 60, 40, 45, 30, 60, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

    output_file = os.path.join(BASE_DIR, 'multi_round_results.xlsx')
    wb.save(output_file)
    return {"success": True, "data": {"output": output_file, "rows": row - 2}}


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Multi-round helper")
    parser.add_argument("--action", required=True,
                        choices=["git_upload", "clean_trace", "fill_dissatisfaction",
                                 "generate_next_prompt", "export_xlsx"])
    parser.add_argument("--tasks-json", default=None, help="JSON string of tasks array")
    parser.add_argument("--tasks-json-file", default=None, help="File containing tasks JSON")
    parser.add_argument("--round", type=int, default=1, help="Current round number")
    parser.add_argument("--rounds", type=int, default=3, help="Total rounds (for export)")
    parser.add_argument("--repo", default=get_config("github.repository", ""))
    parser.add_argument("--branch", default=get_config("github.branch", "main"))
    parser.add_argument("--prefix", default=get_config("github.folder_prefix", "cp1"))
    parser.add_argument("--date", default=None)

    args = parser.parse_args()

    # 优先从文件读取 tasks JSON
    if args.tasks_json_file:
        with open(args.tasks_json_file, 'r', encoding='utf-8') as f:
            args.tasks_json = f.read()
    if not args.tasks_json:
        print(json.dumps({"success": False, "error": "Missing --tasks-json or --tasks-json-file"}))
        sys.exit(1)

    try:
        if args.action == "git_upload":
            result = action_git_upload(args)
        elif args.action == "clean_trace":
            result = action_clean_trace(args)
        elif args.action == "fill_dissatisfaction":
            result = action_fill_dissatisfaction(args)
        elif args.action == "generate_next_prompt":
            result = action_generate_next_prompt(args)
        elif args.action == "export_xlsx":
            result = action_export_xlsx(args)
        else:
            result = {"success": False, "error": f"Unknown action: {args.action}"}
    except Exception as e:
        result = {"success": False, "error": f"Internal error: {e}"}

    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(json.dumps({"success": False, "error": f"Fatal error: {e}"}, ensure_ascii=False))
        sys.exit(0)
