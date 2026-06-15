#!/usr/bin/env python3
"""
读取 trae_results_summary.xlsx，对每一行调用火山引擎API生成"不满意原因"，
结果写回Excel。

依赖安装：
    pip install openpyxl openai
"""

import json
import os
import sys
import time
from config_loader import get_config, get_deepseek_config

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    from openai import OpenAI
except ImportError:
    print("请先安装 openai: pip install openai")
    sys.exit(1)

# ===================== API 配置 =====================
API_KEY, BASE_URL, MODEL = get_deepseek_config()
MAX_RETRIES = int(get_config("deepseek.max_retries", 3))
RETRY_DELAY = int(get_config("deepseek.retry_delay_seconds", 2))
# ====================================================

client = OpenAI(api_key=API_KEY, base_url=BASE_URL)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "trae_results_summary.xlsx")
TASKS_DIR = os.path.join(BASE_DIR, "tasks")

SYSTEM_PROMPT = """你是一位技术评审人员，需要根据用户需求（Prompt）和AI的执行过程记录（Trace），对AI交付的产物进行严格评审，写出具体的"产物不满意"和"过程不满意"反馈。

请严格按照以下格式输出（带编号列表，每条具体明确）：

产物不满意：
<逐条列出产物本身存在的具体问题。不能只写笼统结论，必须说明具体问题点。每条原因应至少包含以下要素中的两项及以上：
- 范围/对象：问题发生在哪个页面、路由、接口、功能模块、文件、命令或运行环境。
- 现象/证据：实际发生了什么，包括报错信息、日志、编译错误、HTTP状态码、操作步骤、复现结果。
- 与需求的偏差：相比需求缺了什么、错在哪里，期望是什么、实际是什么。
- 影响范围：影响了哪些核心流程或功能，如登录、下单、提交、跳转、数据展示、游戏流程、服务启动等。
- 复现条件：在什么条件下出现，如使用的账号、输入数据、浏览器/设备、运行命令、环境版本、特定操作路径。
- 严重程度：问题是阻塞主流程、部分功能异常、体验问题，还是边界场景问题。>

过程不满意：
<逐条列出AI开发过程中的问题。说明做事方式哪里不靠谱，如改了一堆东西却没把核心需求做完、绕来绕去走了弯路、明明报错了还继续往下做、遗留了一堆坑没收尾、核心功能未做闭环验证等。>

硬性要求：
1. 不满意理由不能只写"无法运行、页面打不开、问题没修复、接口失败"等笼统结论，必须说明具体问题点。
2. 产物不满意原因要针对提示词，明确说明是要求中的哪一项完成情况不好，为什么不好：
   - 页面404、空白等，要落实到具体渲染问题、路径问题，不能只写"404"。
   - 接口调用错误，要根据返回的是400还是500，落实到具体前端或后端的代码问题上。
   - 细写出bug原因以及触发方式，不能直接复制大段报错信息。
   - 如果整体项目运行不起来，要解释出具体问题原因：具体的编译错误、依赖冲突、语法问题。
   - 业务逻辑上的问题，要具体说明哪个业务未完成，需要举例说明。
   - 美观度或游戏不好玩等主观感受可以写但不能作为主要评判标准，重点要落实到模型的问题上。
3. 原因模糊不清不够详细（比如"结果不满意有bug"、"项目没跑起来"）直接废弃——必须具体到技术细节。
4. 不能把原因直接复制到下轮Prompt——必须重新组织语言，提炼具体问题。
5. 不要特别口语化，也不要特别像AI写的。语气要像有经验的技术负责人在指出具体问题，客观、直接、有依据。
6. 必须严格按"产物不满意："和"过程不满意："格式开头，各占一段，不要合并成一段。
7. 严禁出现"AI"、"模型"、"agent"、"执行步骤"、"工具调用"等元描述词汇。
"""


def analyze_task(prompt: str, trace: str) -> str | None:
    """调用API分析prompt和trace，返回不满意原因文本，带重试机制。"""
    if not prompt:
        prompt = "（无prompt内容）"
    if not trace:
        trace = "（无trace内容）"

    user_content = f"""【用户需求 Prompt】：
{prompt}

【AI执行过程 Trace】：
{trace}

请根据以上内容，严格按照格式要求写出产物不满意原因和过程不满意原因。
注意：
- 产物不满意原因必须逐条列出，针对提示词中的具体需求项，说明哪一项完成得不好、为什么不好。
- 不能写笼统结论，必须具体到技术细节（如具体页面、接口、报错信息、代码问题）。
- 过程不满意原因要说明做事方式的问题，如核心需求未完成、报错未处理、遗留问题未收尾等。"""

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user_content},
                ],
                temperature=0.3,
                max_tokens=4096,
            )
            content = response.choices[0].message.content
            if content and "产物不满意" in content and "过程不满意" in content:
                return content
            print(f"    返回格式异常，重试 ({attempt}/{MAX_RETRIES})...")
        except Exception as e:
            print(f"    API调用失败 ({attempt}/{MAX_RETRIES}): {e}")

        if attempt < MAX_RETRIES:
            time.sleep(RETRY_DELAY * attempt)

    return None


def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"未找到Excel文件: {EXCEL_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"表头: {headers}")

    try:
        dissatisfaction_col = headers.index("不满意原因") + 1
    except ValueError:
        print("Excel中未找到'不满意原因'列，请确认表头名称")
        sys.exit(1)

    try:
        task_name_col = headers.index("task_name") + 1
    except ValueError:
        print("Excel中未找到'task_name'列")
        sys.exit(1)

    try:
        trace_col = headers.index("trace") + 1
    except ValueError:
        print("Excel中未找到'trace'列，请确认表头名称")
        sys.exit(1)

    total = 0
    success = 0
    skip = 0

    for row in range(2, ws.max_row + 1):
        task_name = ws.cell(row=row, column=task_name_col).value
        current_value = ws.cell(row=row, column=dissatisfaction_col).value

        if not task_name:
            continue

        # 如果已有值，跳过
        if current_value and str(current_value).strip():
            print(f"跳过 {task_name}: 已有不满意原因")
            skip += 1
            continue

        # trace 直接从 Excel 读取，不再使用 JSON 中的 trace（JSON 中常存为字符串 "null"）
        trace = ws.cell(row=row, column=trace_col).value or ""
        trace = str(trace).strip()

        result_file = os.path.join(TASKS_DIR, task_name, f"result_{task_name}.json")
        if not os.path.exists(result_file):
            print(f"跳过 {task_name}: 未找到结果文件 {result_file}")
            continue

        try:
            with open(result_file, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"跳过 {task_name}: 结果文件读取失败 - {e}")
            continue

        prompt = data.get("prompt", "")

        if not prompt and not trace:
            print(f"跳过 {task_name}: prompt和trace均为空")
            continue

        print(f"[{total + 1}] 正在分析 {task_name}...")
        result = analyze_task(prompt, trace)
        if result:
            ws.cell(row=row, column=dissatisfaction_col).value = result
            wb.save(EXCEL_FILE)
            print(f"    ✅ {task_name} 已保存")
            success += 1
        else:
            print(f"    ❌ {task_name} 分析失败")

        total += 1
        # 避免触发API速率限制
        if row < ws.max_row:
            time.sleep(1)

    print(f"\n完成！本次处理 {total} 条，成功 {success} 条，跳过 {skip} 条")
    wb.save(EXCEL_FILE)
    print(f"Excel已保存: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
